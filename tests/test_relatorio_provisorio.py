"""Provisional Excel export for the thesis (no full event list required)."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from export_relatorio_provisorio import (  # noqa: E402
    SHEET_NAMES,
    export_relatorio_provisorio,
)

RESUMO = ROOT / "reports" / "campo_resumo.json"


def test_export_relatorio_provisorio_sem_resultado_json(tmp_path: Path) -> None:
    out = tmp_path / "relatorio_provisorio.xlsx"
    path = export_relatorio_provisorio(
        RESUMO, out, resultado_path=None, spectrogram_dir=tmp_path
    )

    assert path.is_file()
    assert path.stat().st_size > 0

    wb = load_workbook(path)
    assert list(wb.sheetnames) == list(SHEET_NAMES)

    resumo_map = {
        row[0]: row[1]
        for row in wb["Resumo"].iter_rows(min_row=1, max_col=2, values_only=True)
        if row[0]
    }
    assert resumo_map["n_eventos"] == 149962
    assert resumo_map["n_arquivos"] == 75

    ouvir = " ".join(
        str(value)
        for row in wb["Ouvir"].iter_rows(max_col=2, values_only=True)
        for value in row
        if value
    )
    assert "R20241012-041002.WAV" in ouvir
    assert "30:45" in ouvir

    spec = " ".join(
        str(value)
        for row in wb["Espectrogramas"].iter_rows(max_col=4, values_only=True)
        for value in row
        if value
    )
    assert "reports/espectrogramas/" in spec
    assert "ausente" in spec


def test_export_relatorio_embute_png(tmp_path: Path) -> None:
    from PIL import Image

    png = tmp_path / "R20241012-041002_30m45s.png"
    Image.new("RGB", (80, 40), color=(20, 20, 20)).save(png)

    out = tmp_path / "relatorio.xlsx"
    path = export_relatorio_provisorio(
        RESUMO, out, resultado_path=None, spectrogram_dir=tmp_path
    )
    wb = load_workbook(path)
    ws = wb["Espectrogramas"]
    assert ws._images
    table = " ".join(
        str(value)
        for row in ws.iter_rows(min_row=4, max_row=6, max_col=3, values_only=True)
        for value in row
        if value
    )
    assert "embutida" in table

