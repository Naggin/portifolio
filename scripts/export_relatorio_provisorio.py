"""Excel provisório e partilhável para a tese (sem as 149 962 linhas de eventos).

Lê sempre ``reports/campo_resumo.json`` (versionado). Se existir
``output/resultado.json`` (gitignored, ~45 MB), acrescenta na aba Ouvir os
64 eventos do recorte de escuta já escolhido.

Uso::

    python3 scripts/export_relatorio_provisorio.py
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import fields
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from bioacoustics.config import DetectionConfig  # noqa: E402

DEFAULT_RESUMO = ROOT / "reports" / "campo_resumo.json"
DEFAULT_OUT = ROOT / "reports" / "relatorio_provisorio.xlsx"
DEFAULT_RESULTADO = ROOT / "output" / "resultado.json"
DEFAULT_SPEC_DIR = ROOT / "reports" / "espectrogramas"

SHEET_NAMES = (
    "Leia primeiro",
    "Resumo",
    "Campanhas",
    "Arquivos",
    "Por hora",
    "Por mês",
    "Ouvir",
    "Espectrogramas",
    "Parâmetros",
)

# Recorte de validação combinado com a aluna (não é o WAV mais denso).
LISTEN = {
    "campaign": "10_10_25 açude 1",
    "file": "R20241012-041002.WAV",
    "recorded_at": "2024-10-12T04:10:02",
    "start_s": 1845.0,  # 30:45
    "end_s": 1905.0,  # 31:45
    "seek": "30:45",
    "listen_until": "31:45",
    "clock_start": "04:40:47",
    "clock_end": "04:41:47",
    "expected_n": 64,
    "median_duration_s": 0.1,
    "median_freq_hz": 2918.0,
    "reference": "Áudio base/CEAES 2.m4a",
    "reference_peak_hz": 2890.0,
    "avoid_file": "R20241012-091022.WAV",
    "avoid_clock": "09:10",
}

MONTH_LABELS = (
    "jan", "fev", "mar", "abr", "mai", "jun",
    "jul", "ago", "set", "out", "nov", "dez",
)

PARAM_NOTES: dict[str, str] = {
    "sample_rate": "Hz; áudio é desamostrado na leitura.",
    "lowcut_hz": "Limite inferior do passa-faixa (calibração CEAES 2).",
    "highcut_hz": "Limite superior do passa-faixa.",
    "filter_order": "Butterworth + sosfiltfilt (fase zero).",
    "n_fft": "Tamanho da janela STFT.",
    "hop_length": "Salto da STFT (amostras).",
    "use_noise_reduction": "Aplica noisereduce na banda.",
    "noise_reduce_prop_decrease": "Intensidade da redução de ruído.",
    "threshold_k": "Limiar = mediana + k·MAD da energia da banda (por janela de 60 s).",
    "min_peak_distance_s": "Declarado em DetectionConfig; NÃO é usado em detect_events.",
    "event_merge_gap_s": "Fundir segmentos de energia com intervalo ≤ este valor.",
    "min_event_duration_s": "Descarta eventos mais curtos que isto.",
    "freq_separation_hz": "Picos espectrais mais próximos contam como o mesmo cantor.",
    "caller_rel_height": "Fração do pico do quadro para contar um cantor distinto.",
    "edge_guard_s": "Margem morta no início/fim de cada janela (filtro fase zero).",
    "chunk_duration_s": "Duração da janela em arquivos longos.",
    "chunk_overlap_s": "Sobreposição entre janelas (2 × edge_guard_s).",
    "spectrogram_preview_s": "PNG de arquivos longos = janela de 60 s com mais picos, não os primeiros 60 s. Validação da tese: aba Espectrogramas.",
}

BANNER_FILL = PatternFill("solid", fgColor="C65911")
BANNER_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(name="Calibri", italic=True, size=10, color="7F6000")
LABEL_FONT = Font(name="Calibri", bold=True, size=11)
BODY_FONT = Font(name="Calibri", size=11)
PEAK_FILL = PatternFill("solid", fgColor="C6EFCE")
ZERO_FILL = PatternFill("solid", fgColor="F2F2F2")
MISSING_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center")
LEFT_TOP = Alignment(vertical="top", horizontal="left")

BANNER_TEXT = (
    "PROVISÓRIO — eventos acústicos na banda calibrada 2,6–3,2 kHz; "
    "não é censo de machos nem prova de espécie em cada linha."
)


VALIDATION_FIGURES = (
    {
        "file": "R20241012-041002_30m45s.png",
        "caption": (
            "R20241012-041002.WAV — 30:45–31:45 (04:40) — picos de vocalização. "
            "Recorte de escuta já combinado (64 eventos, ~2,93 kHz). "
            "Círculos verdes = eventos do lote, não uma recontagem."
        ),
    },
    {
        "file": "R20241012-041002_pico_1854s.png",
        "caption": (
            "Zoom ~8 s no pico mais forte deste minuto (~1853,65 s, 2928 Hz). "
            "Só para conferir a forma do canto; o método de contagem continua a ser "
            "a matriz numérica da STFT."
        ),
    },
    {
        "file": "R20241012-091022_densest_60s.png",
        "caption": (
            "R20241012-091022.WAV — janela de 60 s mais densa — só depois da madrugada. "
            "Não começar a validação por este ficheiro."
        ),
    },
)


def export_relatorio_provisorio(
    resumo_path: Path,
    out_path: Path,
    resultado_path: Path | None = None,
    spectrogram_dir: Path | None = None,
) -> Path:
    """Write the provisional workbook. ``resultado_path`` is optional."""
    resumo = json.loads(Path(resumo_path).read_text(encoding="utf-8"))
    listen_events = _load_listen_events(resultado_path)
    spec_dir = Path(spectrogram_dir) if spectrogram_dir is not None else DEFAULT_SPEC_DIR

    wb = Workbook()
    ws0 = wb.active
    assert ws0 is not None
    _sheet_leia_primeiro(ws0, resumo, listen_events)
    _sheet_resumo(wb.create_sheet(SHEET_NAMES[1]), resumo)
    _sheet_campanhas(wb.create_sheet(SHEET_NAMES[2]), resumo)
    _sheet_arquivos(wb.create_sheet(SHEET_NAMES[3]), resumo)
    _sheet_por_hora(wb.create_sheet(SHEET_NAMES[4]), resumo)
    _sheet_por_mes(wb.create_sheet(SHEET_NAMES[5]), resumo)
    _sheet_ouvir(wb.create_sheet(SHEET_NAMES[6]), listen_events)
    _sheet_espectrogramas(wb.create_sheet(SHEET_NAMES[7]), spec_dir)
    _sheet_parametros(wb.create_sheet(SHEET_NAMES[8]))

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _load_listen_events(resultado_path: Path | None) -> list[dict[str, Any]] | None:
    if resultado_path is None:
        return None
    path = Path(resultado_path)
    if not path.is_file():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    lo = float(LISTEN["start_s"])
    hi = float(LISTEN["end_s"])
    fname = str(LISTEN["file"])
    rows = [
        event
        for event in payload.get("events", [])
        if event.get("file") == fname
        and lo <= float(event["peak_time_s"]) < hi
    ]
    rows.sort(key=lambda event: float(event["peak_time_s"]))
    return rows


def _banner(ws: Worksheet, n_cols: int, text: str = BANNER_TEXT) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_cols, 2))
    cell = ws.cell(1, 1, text)
    cell.fill = BANNER_FILL
    cell.font = BANNER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 36
    for col in range(1, n_cols + 1):
        ws.cell(1, col).fill = BANNER_FILL


def _note(ws: Worksheet, row: int, n_cols: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(n_cols, 2))
    cell = ws.cell(row, 1, text)
    cell.fill = NOTE_FILL
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 48
    for col in range(1, n_cols + 1):
        ws.cell(row, col).fill = NOTE_FILL


def _headers(ws: Worksheet, row: int, titles: list[str]) -> None:
    for col, title in enumerate(titles, start=1):
        cell = ws.cell(row, col, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
        cell.border = THIN
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = f"A{row + 1}"


def _cell(ws: Worksheet, row: int, col: int, value: Any, fmt: str | None = None) -> None:
    cell = ws.cell(row, col, value)
    cell.font = BODY_FONT
    cell.alignment = LEFT_TOP
    cell.border = THIN
    if fmt:
        cell.number_format = fmt


def _widths(ws: Worksheet, widths: list[float]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _tab(ws: Worksheet, color: str) -> None:
    ws.sheet_properties.tabColor = color


def _clock(recorded_at: str, peak_time_s: float) -> str:
    stamp = datetime.fromisoformat(recorded_at)
    return (stamp + timedelta(seconds=float(peak_time_s))).strftime("%H:%M:%S")


def _bar_chart(
    ws: Worksheet,
    title: str,
    x_title: str,
    header_row: int,
    n_rows: int,
    data_col: int = 2,
    cat_col: int = 1,
    anchor: str = "E3",
) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.x_axis.title = x_title
    chart.y_axis.title = "Eventos"
    chart.style = 10
    chart.legend = None
    chart.y_axis.numFmt = "#,##0"
    data = Reference(ws, min_col=data_col, min_row=header_row, max_row=header_row + n_rows)
    cats = Reference(ws, min_col=cat_col, min_row=header_row + 1, max_row=header_row + n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, anchor)


def _sheet_leia_primeiro(
    ws: Worksheet,
    resumo: dict[str, Any],
    listen_events: list[dict[str, Any]] | None,
) -> None:
    ws.title = SHEET_NAMES[0]
    _tab(ws, "C65911")
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    n_listen = len(listen_events) if listen_events is not None else None
    listen_status = (
        f"{n_listen} eventos do recorte listados na aba Ouvir "
        "(extraídos de output/resultado.json)."
        if n_listen is not None
        else (
            "Lista dos 64 eventos do recorte ausente: rode o script na máquina "
            "com output/resultado.json para a preencher."
        )
    )
    pipeline = resumo.get("pipeline_generated_at") or "lote de campo"
    blocks = [
        (
            "Estado",
            "Este livro é PROVISÓRIO: serve para a aluna abrir, enviar por e-mail "
            "e usar no rascunho da tese. Não substitui a lista completa de eventos "
            "(output/resultado.xlsx, ~7 MB, gitignored). A aba Espectrogramas traz "
            "o recorte de escuta da madrugada com os picos marcados (validação "
            "humana). Não há um PNG por evento do lote (isso seriam ~150 mil imagens).",
        ),
        (
            "O que os números são",
            "Contagem automática de eventos acústicos — picos de energia na banda "
            "calibrada 2,6–3,2 kHz (referência Áudio base/CEAES 2.m4a, pico ~2,89 kHz). "
            f"Mesmo DetectionConfig em {resumo['summary']['n_files']} arquivos, "
            f"{resumo['summary']['total_duration_h']} h, "
            f"{resumo['summary']['n_events']} eventos, "
            f"{resumo['summary']['n_errors']} erros de processamento. "
            f"Pipeline gerado em {pipeline}.",
        ),
        (
            "O que NÃO afirmar",
            "Não são 149 962 machos, indivíduos ou pererecas. Um macho gera muitos "
            "eventos; dois machos no mesmo pitch podem virar um. Qualquer energia na "
            "banda acima do limiar vira evento (inseto, gravador, outra espécie). "
            "Não se comprovou espécie até de manhã. max_simultaneous = 2 é teto "
            "estrutural da banda de 600 Hz com separação de 400 Hz, não um censo. "
            "A Fase 3 (Claude) não foi usada.",
        ),
        (
            "Gráficos hora / mês",
            "A hora/mês é a do INÍCIO do arquivo (nome RYYYYMMDD-HHMMSS), não o "
            "relógio de cada canto. Sete MP3s (campanhas 12_09_25 e 15_08_25) não "
            "têm data no nome: 50 170 eventos entram no total e saem dos gráficos. "
            "Zeros às 11–13 h = nenhum WAV a começar nessa hora, não «zero canto». "
            "Agosto/setembro vazios nos gráficos pelas mesmas MP3s.",
        ),
        (
            "Próximo passo",
            "Escuta humana do recorte já escolhido (aba Ouvir): pasta "
            f"«{LISTEN['campaign']}», ficheiro {LISTEN['file']}, "
            f"gravado em {LISTEN['recorded_at']}. Seek {LISTEN['seek']} "
            f"({int(LISTEN['start_s'])} s), ouvir 60 s até {LISTEN['listen_until']}. "
            f"Relógio de campo {LISTEN['clock_start']}–{LISTEN['clock_end']}. "
            f"O detector marcou {LISTEN['expected_n']} eventos curtos (~"
            f"{LISTEN['median_duration_s']} s), mediana ~{LISTEN['median_freq_hz']:.0f} Hz "
            f"(CEAES 2 ~{LISTEN['reference_peak_hz'] / 1000:.2f} kHz). "
            f"Comparar com {LISTEN['reference']}. "
            f"NÃO começar pelo ficheiro da manhã {LISTEN['avoid_file']} "
            f"({LISTEN['avoid_clock']}, o mais denso); só depois de o recorte da "
            "madrugada soar a rã.",
        ),
        (
            "Documentos",
            "O que a tese pode afirmar: docs/PRECISAO_E_LIMITES.md. "
            "Como ler Excel/JSON completos: docs/COMO_LER_OS_RESULTADOS.md. "
            "Parâmetros: aba Parâmetros e src/bioacoustics/config.py. "
            "Totais enxutos: reports/campo_resumo.json. "
            "PNG de validação: aba Espectrogramas e reports/espectrogramas/.",
        ),
        (
            "Como regenerar",
            "python3 scripts/export_relatorio_provisorio.py  "
            f"(gerado {generated}). {listen_status}",
        ),
    ]
    _banner(ws, 2)
    _headers(ws, 2, ["Secção", "Texto"])
    for i, (section, text) in enumerate(blocks):
        row = 3 + i
        ws.cell(row, 1, section).font = LABEL_FONT
        ws.cell(row, 1).alignment = WRAP_TOP
        ws.cell(row, 1).border = THIN
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="D6DCE4")
        body = ws.cell(row, 2, text)
        body.font = BODY_FONT
        body.alignment = WRAP_TOP
        body.border = THIN
        ws.row_dimensions[row].height = 78
    _widths(ws, [22, 110])
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def _sheet_resumo(ws: Worksheet, resumo: dict[str, Any]) -> None:
    _tab(ws, "1F4E79")
    summary = resumo["summary"]
    wav = resumo.get("wav_events_per_hour") or {}
    rows = [
        ("Espécie", resumo.get("species"), "nome científico"),
        ("Nome comum", resumo.get("common_name"), ""),
        ("n_arquivos", summary["n_files"], "gravações de açude 1 (pastas fora de propósito excluídas)"),
        ("Duração total (h)", summary["total_duration_h"], f"{summary['total_duration_s']} s"),
        ("n_eventos", summary["n_events"], "eventos acústicos na banda; NÃO são indivíduos"),
        ("Eventos / h (lote)", summary.get("events_per_hour"), "média global"),
        ("max_simultaneous", summary["max_simultaneous"], "teto estrutural da banda + separação 400 Hz"),
        ("n_erros", summary["n_errors"], "falhas de leitura/processamento"),
        (
            "Mediana eventos/h (WAV)",
            wav.get("median"),
            f"n={wav.get('n')}; mín. {wav.get('min')}; máx. {wav.get('max')}",
        ),
        (
            "MP3s sem recorded_at",
            summary["n_files_without_recorded_at"],
            "campanhas 12_09_25 e 15_08_25; nomes 111009_001 / Z0000002",
        ),
        (
            "Eventos fora dos gráficos hora/mês",
            summary["n_events_without_recorded_at"],
            f"{summary['n_events_in_by_hour_by_month']} eventos entram em Por hora / Por mês",
        ),
        ("Banda (Hz)", f"{resumo['config']['lowcut_hz']:.0f}–{resumo['config']['highcut_hz']:.0f}", f"threshold_k={resumo['config']['threshold_k']}"),
    ]
    _banner(ws, 3)
    _note(
        ws,
        2,
        3,
        "Totais do lote de campo (reports/campo_resumo.json). "
        "Cite como eventos na banda calibrada. Ver docs/PRECISAO_E_LIMITES.md.",
    )
    _headers(ws, 3, ["Indicador", "Valor", "Nota"])
    int_keys = {
        "n_arquivos",
        "n_eventos",
        "max_simultaneous",
        "n_erros",
        "MP3s sem recorded_at",
        "Eventos fora dos gráficos hora/mês",
    }
    for i, (label, value, note) in enumerate(rows):
        row = 4 + i
        _cell(ws, row, 1, label)
        ws.cell(row, 1).font = LABEL_FONT
        _cell(ws, row, 2, value)
        if label in int_keys and isinstance(value, int):
            ws.cell(row, 2).number_format = "#,##0"
        elif isinstance(value, float):
            ws.cell(row, 2).number_format = "0.0"
        _cell(ws, row, 3, note)
        ws.cell(row, 3).alignment = WRAP_TOP
        ws.row_dimensions[row].height = 22
    caveats = resumo.get("caveats") or []
    start = 4 + len(rows) + 1
    ws.cell(start, 1, "Ressalvas do lote").font = LABEL_FONT
    for i, caveat in enumerate(caveats):
        cell = ws.cell(start + 1 + i, 1, f"• {caveat}")
        ws.merge_cells(start_row=start + 1 + i, start_column=1, end_row=start + 1 + i, end_column=3)
        cell.alignment = WRAP_TOP
        cell.font = BODY_FONT
        ws.row_dimensions[start + 1 + i].height = 36
    _widths(ws, [38, 22, 78])


def _sheet_campanhas(ws: Worksheet, resumo: dict[str, Any]) -> None:
    _tab(ws, "2E75B6")
    titles = [
        "Campanha",
        "Arquivos",
        "Eventos",
        "Duração (h)",
        "Eventos/h",
        "Máx. simultâneos",
        "Sem data no nome",
    ]
    _banner(ws, len(titles))
    _note(
        ws,
        2,
        len(titles),
        "Pastas de açude 1. «15/08 açude 2 (esse nao precisa)» e «Áudio base» "
        "não entram. Campanhas só com MP3s não têm recorded_at.",
    )
    _headers(ws, 3, titles)
    for i, camp in enumerate(resumo.get("campaigns") or []):
        row = 4 + i
        values = [
            camp["campaign"],
            camp["n_files"],
            camp["n_events"],
            camp["total_duration_h"],
            camp["events_per_hour"],
            camp["max_simultaneous"],
            camp["n_without_recorded_at"],
        ]
        for col, value in enumerate(values, start=1):
            _cell(ws, row, col, value)
        ws.cell(row, 2).number_format = "0"
        ws.cell(row, 3).number_format = "#,##0"
        ws.cell(row, 4).number_format = "0.000"
        ws.cell(row, 5).number_format = "0.0"
        if camp.get("n_without_recorded_at"):
            for col in range(1, len(titles) + 1):
                ws.cell(row, col).fill = MISSING_FILL
    _widths(ws, [24, 12, 14, 14, 14, 18, 18])


def _sheet_arquivos(ws: Worksheet, resumo: dict[str, Any]) -> None:
    _tab(ws, "2E75B6")
    titles = [
        "Campanha",
        "Arquivo",
        "Gravado em",
        "Duração (s)",
        "Eventos",
        "Eventos/h",
        "Máx. simultâneos",
    ]
    _banner(ws, len(titles))
    _note(
        ws,
        2,
        len(titles),
        "Uma linha por gravação. Gravado em vem do nome RYYYYMMDD-HHMMSS; "
        "vazio nos 7 MP3s. O ficheiro de escuta é R20241012-041002.WAV "
        "(não R20241012-091022.WAV, o mais denso).",
    )
    _headers(ws, 3, titles)
    listen_file = str(LISTEN["file"])
    avoid_file = str(LISTEN["avoid_file"])
    for i, item in enumerate(resumo.get("files") or []):
        row = 4 + i
        recorded = item.get("recorded_at") or ""
        values = [
            item.get("campaign"),
            item.get("file"),
            recorded,
            item.get("duration_s"),
            item.get("n_events"),
            item.get("events_per_hour"),
            item.get("max_simultaneous"),
        ]
        for col, value in enumerate(values, start=1):
            _cell(ws, row, col, value)
        ws.cell(row, 4).number_format = "0.0"
        ws.cell(row, 5).number_format = "#,##0"
        ws.cell(row, 6).number_format = "0.0"
        name = item.get("file")
        if name == listen_file:
            for col in range(1, len(titles) + 1):
                ws.cell(row, col).fill = PEAK_FILL
        elif name == avoid_file or not recorded:
            for col in range(1, len(titles) + 1):
                ws.cell(row, col).fill = MISSING_FILL if not recorded else ZERO_FILL
    last = 3 + len(resumo.get("files") or [])
    if last >= 3:
        ws.auto_filter.ref = f"A3:G{last}"
    _widths(ws, [22, 28, 22, 14, 12, 14, 18])


def _sheet_por_hora(ws: Worksheet, resumo: dict[str, Any]) -> None:
    _tab(ws, "548235")
    titles = ["Hora", "Eventos", "Nota"]
    _banner(ws, 6)
    _note(
        ws,
        2,
        6,
        "Hora = INÍCIO do arquivo, não o instante de cada evento. "
        "Zeros às 11–13 h: nenhum WAV a começar então. Pico às 4–5 h. "
        "50 170 eventos de MP3 sem data não entram aqui.",
    )
    _headers(ws, 3, titles)
    peak_hours = {4, 5}
    zero_hours = {11, 12, 13}
    by_hour = resumo.get("by_hour") or [{"hour": h, "n_events": 0} for h in range(24)]
    for i, item in enumerate(by_hour):
        row = 4 + i
        hour = int(item["hour"])
        n_events = int(item["n_events"])
        label = f"{hour:02d}h"
        if hour in peak_hours:
            note = "pico (início de arquivo neste bloco)"
        elif hour in zero_hours:
            note = "nenhum WAV a começar nesta hora (não é «zero canto»)"
        elif n_events == 0:
            note = "sem arquivos com recorded_at nesta hora"
        else:
            note = ""
        _cell(ws, row, 1, label)
        _cell(ws, row, 2, n_events, "#,##0")
        _cell(ws, row, 3, note)
        fill = PEAK_FILL if hour in peak_hours else ZERO_FILL if hour in zero_hours else None
        if fill:
            for col in range(1, 4):
                ws.cell(row, col).fill = fill
    _bar_chart(
        ws,
        "Eventos por hora (início do arquivo)",
        "Hora (início do arquivo)",
        header_row=3,
        n_rows=24,
        anchor="E3",
    )
    _widths(ws, [10, 14, 62, 14, 14, 14])


def _sheet_por_mes(ws: Worksheet, resumo: dict[str, Any]) -> None:
    _tab(ws, "548235")
    titles = ["Mês", "Eventos", "Nota"]
    _banner(ws, 6)
    _note(
        ws,
        2,
        6,
        "Mês = recorded_at do INÍCIO do arquivo. Agosto e setembro ficam a zero "
        "porque essas campanhas são MP3s sem data no nome (os eventos existem no "
        "total, aba Resumo). Outubro concentra os WAV datados.",
    )
    _headers(ws, 3, titles)
    by_month = resumo.get("by_month") or [{"month": m, "n_events": 0} for m in range(1, 13)]
    for i, item in enumerate(by_month):
        row = 4 + i
        month = int(item["month"])
        n_events = int(item["n_events"])
        label = f"{month:02d} {MONTH_LABELS[month - 1]}"
        if month in (8, 9):
            note = "campanhas MP3 sem data no nome — eventos não entram neste gráfico"
            fill = MISSING_FILL
        elif n_events == 0:
            note = "sem WAV com recorded_at neste mês"
            fill = ZERO_FILL
        elif month == 10:
            note = "pico entre os arquivos com data no nome"
            fill = PEAK_FILL
        else:
            note = ""
            fill = None
        _cell(ws, row, 1, label)
        _cell(ws, row, 2, n_events, "#,##0")
        _cell(ws, row, 3, note)
        if fill:
            for col in range(1, 4):
                ws.cell(row, col).fill = fill
    _bar_chart(
        ws,
        "Eventos por mês (início do arquivo)",
        "Mês (início do arquivo)",
        header_row=3,
        n_rows=12,
        anchor="E3",
    )
    _widths(ws, [14, 14, 72, 14, 14, 14])


def _sheet_ouvir(ws: Worksheet, listen_events: list[dict[str, Any]] | None) -> None:
    _tab(ws, "C65911")
    info = [
        ("Pasta", LISTEN["campaign"]),
        ("Arquivo", LISTEN["file"]),
        ("Gravado em", LISTEN["recorded_at"]),
        ("Seek no leitor", f"{LISTEN['seek']}  ({int(LISTEN['start_s'])} s)"),
        ("Ouvir até", f"{LISTEN['listen_until']}  ({int(LISTEN['end_s'])} s) — 60 s"),
        ("Relógio de campo", f"{LISTEN['clock_start']}–{LISTEN['clock_end']}"),
        (
            "O que o detector marcou",
            f"{LISTEN['expected_n']} eventos, curtos (~{LISTEN['median_duration_s']} s), "
            f"mediana ~{LISTEN['median_freq_hz']:.0f} Hz "
            f"(CEAES 2 pico ~{LISTEN['reference_peak_hz'] / 1000:.2f} kHz)",
        ),
        ("Comparar com", LISTEN["reference"]),
        (
            "Não começar por",
            f"{LISTEN['avoid_file']} ({LISTEN['avoid_clock']}, o WAV mais denso). "
            "Só depois de este recorte da madrugada soar a rã.",
        ),
        (
            "Como ouvir",
            "Abrir o WAV no VLC/Audacity, ir a 30:45, ouvir um minuto com fones. "
            "A pergunta é: soa como o anúncio do CEAES 2, ou é outro som na banda? "
            "O espectrograma deste minuto (picos marcados) está na aba Espectrogramas.",
        ),
    ]
    titles = [
        "#",
        "peak_time_s",
        "Relógio",
        "peak_freq_hz",
        "duration_s",
        "energy",
        "n_callers",
    ]
    n_cols = len(titles)
    _banner(ws, n_cols)
    _note(
        ws,
        2,
        n_cols,
        "Recorte de validação humana (já escolhido). PROVISÓRIO: a escuta ainda "
        "não foi feita neste lote. Confirmar espécie no ouvido antes de discutir "
        "a manhã densa. Espectrograma deste minuto: aba Espectrogramas.",
    )
    for i, (label, value) in enumerate(info):
        row = 3 + i
        ws.cell(row, 1, label).font = LABEL_FONT
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="D6DCE4")
        ws.cell(row, 1).alignment = WRAP_TOP
        ws.cell(row, 1).border = THIN
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n_cols)
        body = ws.cell(row, 2, value)
        body.font = BODY_FONT
        body.alignment = WRAP_TOP
        body.border = THIN
        ws.row_dimensions[row].height = 22

    header_row = 3 + len(info) + 1
    _headers(ws, header_row, titles)
    # _headers sets freeze to header+1; keep the instruction block visible too.
    ws.freeze_panes = f"A{header_row + 1}"

    if listen_events is None:
        row = header_row + 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        cell = ws.cell(
            row,
            1,
            "Lista dos 64 eventos indisponível neste ficheiro: "
            "output/resultado.json não estava presente (CI / máquina sem o JSON "
            "completo). Na cópia da tese gerada com o JSON local, esta tabela "
            "tem uma linha por evento (peak_time_s, relógio, frequência, duração, "
            "energia, n_callers). O recorte em si não muda.",
        )
        cell.alignment = WRAP_TOP
        cell.fill = NOTE_FILL
        cell.font = NOTE_FONT
        ws.row_dimensions[row].height = 48
    else:
        recorded = str(LISTEN["recorded_at"])
        for i, event in enumerate(listen_events, start=1):
            row = header_row + i
            peak = float(event["peak_time_s"])
            values = [
                i,
                round(peak, 3),
                _clock(str(event.get("recorded_at") or recorded), peak),
                round(float(event["peak_freq_hz"]), 1),
                round(float(event["duration_s"]), 3),
                round(float(event["energy"]), 3),
                int(event["n_callers"]),
            ]
            for col, value in enumerate(values, start=1):
                _cell(ws, row, col, value)
            ws.cell(row, 2).number_format = "0.000"
            ws.cell(row, 4).number_format = "0.0"
            ws.cell(row, 5).number_format = "0.000"
            ws.cell(row, 6).number_format = "0.000"
        n = len(listen_events)
        summary_row = header_row + n + 1
        if n:
            freqs = [float(event["peak_freq_hz"]) for event in listen_events]
            durs = [float(event["duration_s"]) for event in listen_events]
            extra = (
                f"Duração mediana {sorted(durs)[len(durs) // 2]:.3f} s; "
                f"frequência mediana {sorted(freqs)[len(freqs) // 2]:.1f} Hz."
            )
        else:
            extra = "Nenhum evento neste intervalo no JSON."
        ws.cell(
            summary_row,
            1,
            (
                f"{n} eventos no recorte {LISTEN['seek']}–{LISTEN['listen_until']} "
                f"(ficheiro {LISTEN['file']}). {extra}"
            ),
        ).font = NOTE_FONT
        ws.merge_cells(
            start_row=summary_row, start_column=1, end_row=summary_row, end_column=n_cols
        )
    _widths(ws, [28, 16, 14, 16, 14, 12, 12])


def _sheet_espectrogramas(ws: Worksheet, spectrogram_dir: Path) -> None:
    _tab(ws, "833C0C")
    n_cols = 4
    _banner(ws, n_cols)
    _note(
        ws,
        2,
        n_cols,
        "Validação humana, não o método de contagem. O lote de campo não gerou "
        "um PNG por evento (seriam 149 962 imagens). Aqui só o recorte de escuta "
        "da madrugada (30:45) com os picos já encontrados marcados em verde, "
        "mais um zoom no pico mais forte. Em arquivos longos o PNG do pipeline "
        "passa a ser a janela de 60 s com mais picos, não os primeiros 60 s. "
        "Ficheiros: reports/espectrogramas/.",
    )
    titles = ["Figura", "Ficheiro", "Estado", "Legenda"]
    _headers(ws, 3, titles)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 88

    img_row = 4 + len(VALIDATION_FIGURES) + 2
    for i, fig in enumerate(VALIDATION_FIGURES):
        row = 4 + i
        png = Path(spectrogram_dir) / str(fig["file"])
        present = png.is_file()
        _cell(ws, row, 1, i + 1)
        _cell(ws, row, 2, f"reports/espectrogramas/{fig['file']}")
        _cell(ws, row, 3, "embutida" if present else "ausente")
        _cell(ws, row, 4, fig["caption"])
        ws.cell(row, 4).alignment = WRAP_TOP
        ws.row_dimensions[row].height = 36
        if not present:
            ws.cell(row, 3).fill = MISSING_FILL
            continue
        caption_row = img_row
        ws.merge_cells(start_row=caption_row, start_column=1, end_row=caption_row, end_column=n_cols)
        cap = ws.cell(caption_row, 1, fig["caption"])
        cap.font = LABEL_FONT
        cap.alignment = WRAP_TOP
        ws.row_dimensions[caption_row].height = 32
        image_row = caption_row + 1
        try:
            img = XLImage(str(png))
            orig_w = float(img.width or 1200)
            orig_h = float(img.height or 650)
            max_w = 640.0
            ratio = min(1.0, max_w / orig_w)
            img.width = int(orig_w * ratio)
            img.height = int(orig_h * ratio)
            height_pt = max(200, int(img.height * 0.78))
            ws.row_dimensions[image_row].height = height_pt
            ws.add_image(img, f"A{image_row}")
        except Exception:
            miss = ws.cell(image_row, 1, f"Falha ao embutir {png.name}")
            miss.fill = MISSING_FILL
            miss.font = NOTE_FONT
        img_row = image_row + 2

    hint_row = 4 + len(VALIDATION_FIGURES)
    ws.merge_cells(start_row=hint_row, start_column=1, end_row=hint_row, end_column=n_cols)
    hint = ws.cell(
        hint_row,
        1,
        "Regenerar PNGs (máquina com o WAV de campo e output/resultado.json): "
        "python3 scripts/export_espectrogramas_validacao.py  "
        "depois  python3 scripts/export_relatorio_provisorio.py",
    )
    hint.font = NOTE_FONT
    hint.fill = NOTE_FILL
    hint.alignment = WRAP_TOP
    ws.row_dimensions[hint_row].height = 36


def _sheet_parametros(ws: Worksheet) -> None:
    _tab(ws, "833C0C")
    titles = ["Parâmetro", "Valor padrão", "Nota"]
    _banner(ws, 3)
    _note(
        ws,
        2,
        3,
        "Padrões de DetectionConfig em src/bioacoustics/config.py. "
        "Este relatório NÃO altera o detector. Calibração: CEAES 2 (Fase 2).",
    )
    _headers(ws, 3, titles)
    cfg = DetectionConfig()
    for i, field in enumerate(fields(DetectionConfig)):
        row = 4 + i
        value = getattr(cfg, field.name)
        _cell(ws, row, 1, field.name)
        ws.cell(row, 1).font = Font(name="Consolas", size=11)
        display: Any = str(value) if isinstance(value, bool) else value
        _cell(ws, row, 2, display)
        _cell(ws, row, 3, PARAM_NOTES.get(field.name, ""))
        ws.cell(row, 3).alignment = WRAP_TOP
        if field.name == "min_peak_distance_s":
            for col in range(1, 4):
                ws.cell(row, col).fill = MISSING_FILL
        ws.row_dimensions[row].height = 20
    _widths(ws, [28, 16, 78])


def main(argv: list[str] | None = None) -> Path:
    parser = argparse.ArgumentParser(
        description="Gera reports/relatorio_provisorio.xlsx a partir de campo_resumo.json.",
    )
    parser.add_argument("--resumo", type=Path, default=DEFAULT_RESUMO)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument(
        "--resultado",
        type=Path,
        default=None,
        help="JSON completo do lote (opcional). Se omitido, usa output/resultado.json se existir.",
    )
    parser.add_argument(
        "--sem-resultado",
        action="store_true",
        help="Não ler output/resultado.json (útil em testes/CI).",
    )
    parser.add_argument(
        "--espectrogramas",
        type=Path,
        default=DEFAULT_SPEC_DIR,
        help="Pasta com os PNG de validação a embutir na aba Espectrogramas.",
    )
    args = parser.parse_args(argv)

    resultado: Path | None
    if args.sem_resultado:
        resultado = None
    elif args.resultado is not None:
        resultado = args.resultado
    elif DEFAULT_RESULTADO.is_file():
        resultado = DEFAULT_RESULTADO
    else:
        resultado = None

    path = export_relatorio_provisorio(
        args.resumo, args.out, resultado, spectrogram_dir=args.espectrogramas
    )
    size_kb = path.stat().st_size / 1024
    extra = "com recorte Ouvir" if resultado else "sem resultado.json"
    print(f"wrote {path} ({size_kb:.1f} KiB, {extra})")
    return path


if __name__ == "__main__":
    main()
