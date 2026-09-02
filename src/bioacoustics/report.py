"""Excel and JSON report generation with per-event detail and hourly/monthly aggregates."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import DetectionConfig
from .pipeline import PipelineResult

SPECIES_NAME = "Sphaenorhynchus caramaschii"
SPECIES_COMMON_NAME = "perereca-de-banhado"
ENERGY_SERIES_POINTS = 400

MONTH_LABELS = (
    "jan", "fev", "mar", "abr", "mai", "jun",
    "jul", "ago", "set", "out", "nov", "dez",
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(name="Calibri", italic=True, size=10, color="7F6000")
BODY_FONT = Font(name="Calibri", size=11)
PEAK_FILL = PatternFill("solid", fgColor="C6EFCE")
ZERO_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _hourly_counts(results: list[PipelineResult]) -> list[dict[str, int]]:
    counts: dict[int, int] = {h: 0 for h in range(24)}
    for r in results:
        if r.recorded_at is None:
            continue
        counts[r.recorded_at.hour] += r.detection.n_events
    return [{"hour": h, "n_events": counts[h]} for h in range(24)]


def _monthly_counts(results: list[PipelineResult]) -> list[dict[str, int]]:
    counts: dict[int, int] = {m: 0 for m in range(1, 13)}
    for r in results:
        if r.recorded_at is None:
            continue
        counts[r.recorded_at.month] += r.detection.n_events
    return [{"month": m, "n_events": counts[m]} for m in range(1, 13)]


def _downsample_series(values: list[float], times: list[float], max_points: int) -> dict[str, list[float]]:
    if len(values) <= max_points:
        return {"times_s": [round(t, 3) for t in times], "energy": [round(v, 4) for v in values]}
    step = max(1, len(values) // max_points)
    idx = list(range(0, len(values), step))
    if idx[-1] != len(values) - 1:
        idx.append(len(values) - 1)
    return {
        "times_s": [round(times[i], 3) for i in idx],
        "energy": [round(values[i], 4) for i in idx],
    }


def _file_payload(result: PipelineResult) -> dict[str, Any]:
    recorded = result.recorded_at.isoformat() if result.recorded_at else None
    times = result.detection.times.tolist()
    energy = result.detection.band_energy.tolist()
    return {
        "file": result.filename,
        "recorded_at": recorded,
        "duration_s": round(result.detection.duration_s, 3),
        "n_events": result.detection.n_events,
        "max_simultaneous": result.detection.max_simultaneous,
        "threshold": round(float(result.detection.threshold), 6),
        "spectrogram": f"{Path(result.filename).stem}_spectrogram.png",
        "band_energy": _downsample_series(energy, times, ENERGY_SERIES_POINTS),
    }


def _event_payloads(result: PipelineResult) -> list[dict[str, Any]]:
    recorded = result.recorded_at.isoformat() if result.recorded_at else None
    rows: list[dict[str, Any]] = []
    for i, ev in enumerate(result.detection.events, start=1):
        rows.append({
            "file": result.filename,
            "recorded_at": recorded,
            "event": i,
            "start_s": round(ev.start_s, 3),
            "end_s": round(ev.end_s, 3),
            "peak_time_s": round(ev.peak_time_s, 3),
            "peak_freq_hz": round(ev.peak_freq_hz, 1),
            "energy": round(ev.energy, 3),
            "n_callers": ev.n_callers,
            "duration_s": round(ev.duration_s, 3),
        })
    return rows


def build_report_payload(
    results: list[PipelineResult],
    cfg: DetectionConfig | None = None,
) -> dict[str, Any]:
    """Serialize pipeline results into the dashboard JSON contract."""
    cfg = cfg or DetectionConfig()
    files = [_file_payload(r) for r in results]
    events = [row for r in results for row in _event_payloads(r)]
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "species": SPECIES_NAME,
        "common_name": SPECIES_COMMON_NAME,
        "config": {
            "sample_rate": cfg.sample_rate,
            "lowcut_hz": cfg.lowcut_hz,
            "highcut_hz": cfg.highcut_hz,
            "threshold_k": cfg.threshold_k,
        },
        "summary": {
            "n_files": len(results),
            "n_events": sum(r.detection.n_events for r in results),
            "max_simultaneous": max((r.detection.max_simultaneous for r in results), default=0),
            "total_duration_s": round(sum(r.detection.duration_s for r in results), 3),
        },
        "files": files,
        "events": events,
        "by_hour": _hourly_counts(results),
        "by_month": _monthly_counts(results),
    }


def write_json_report(
    results: list[PipelineResult],
    out_path: str | Path,
    cfg: DetectionConfig | None = None,
) -> Path:
    """Write the dashboard JSON report next to the Excel workbook."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = build_report_payload(results, cfg)
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return out_path


def write_report(results: list[PipelineResult], out_path: str | Path) -> Path:
    """Write a multi-sheet .xlsx summarizing detections across files."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    _write_events_sheet(wb.active, results)
    _write_files_sheet(wb.create_sheet("Arquivos"), results)
    _write_hourly_sheet(wb.create_sheet("Por hora"), results)
    _write_monthly_sheet(wb.create_sheet("Por mês"), results)

    wb.save(out_path)
    return out_path


def _note_row(ws, row: int, n_cols: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row, 1, text)
    cell.fill = NOTE_FILL
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 40


def _headers(ws, row: int, titles: list[str]) -> None:
    for col, title in enumerate(titles, start=1):
        cell = ws.cell(row, col, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    ws.freeze_panes = f"A{row + 1}"


def _bar_chart(
    ws,
    title: str,
    x_title: str,
    header_row: int,
    n_rows: int,
    anchor: str = "E3",
) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.x_axis.title = x_title
    chart.y_axis.title = "Eventos"
    chart.style = 10
    chart.legend = None
    chart.y_axis.numFmt = "#,##0"
    data = Reference(ws, min_col=2, min_row=header_row, max_row=header_row + n_rows)
    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 10
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "2F5D50"
        d_lbls = DataLabelList()
        d_lbls.showVal = True
        d_lbls.numFmt = "#,##0"
        chart.series[0].dLbls = d_lbls
    ws.add_chart(chart, anchor)


def _write_events_sheet(ws, results: list[PipelineResult]) -> None:
    ws.title = "Eventos"
    titles = [
        "Arquivo", "Gravado em", "Nº", "Início (s)", "Fim (s)",
        "Pico (s)", "Freq. pico (Hz)", "Energia", "Indiv. (est.)",
    ]
    _note_row(
        ws,
        1,
        len(titles),
        "Eventos acústicos na banda calibrada — não são machos. "
        "Pico = instante do pico; Indiv. = picos simultâneos estimados (máx. ~2), não ID.",
    )
    _headers(ws, 2, titles)
    row = 3
    for r in results:
        recorded = r.recorded_at.isoformat() if r.recorded_at else ""
        for i, ev in enumerate(r.detection.events, start=1):
            ws.append([
                r.filename, recorded, i,
                round(ev.start_s, 3), round(ev.end_s, 3),
                round(ev.peak_time_s, 3), round(ev.peak_freq_hz, 1),
                round(ev.energy, 3), ev.n_callers,
            ])
            row += 1
    for col, width in zip(range(1, len(titles) + 1), [28, 22, 6, 12, 12, 12, 14, 10, 12]):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_files_sheet(ws, results: list[PipelineResult]) -> None:
    titles = [
        "Arquivo", "Gravado em", "Duração (s)", "Eventos", "Máx. simultâneos",
    ]
    _note_row(
        ws,
        1,
        len(titles),
        "Uma linha por gravação. Eventos = picos na banda; Máx. simultâneos = teto estrutural (~2).",
    )
    _headers(ws, 2, titles)
    for r in results:
        recorded = r.recorded_at.isoformat() if r.recorded_at else ""
        ws.append([
            r.filename, recorded, round(r.detection.duration_s, 1),
            r.detection.n_events, r.detection.max_simultaneous,
        ])


def _write_hourly_sheet(ws, results: list[PipelineResult]) -> None:
    titles = ["Hora", "Eventos", "Nota"]
    _note_row(
        ws,
        1,
        3,
        "Hora = INÍCIO do arquivo (nome RYYYYMMDD-HHMMSS), não o relógio de cada canto. "
        "Arquivos sem data no nome não entram. Zeros às 11–13 h = nenhum WAV a começar então.",
    )
    _headers(ws, 2, titles)
    peak_hours = {4, 5}
    zero_hours = {11, 12, 13}
    for item in _hourly_counts(results):
        hour = int(item["hour"])
        n_events = int(item["n_events"])
        if hour in peak_hours:
            note = "pico (início de arquivo neste bloco)"
        elif hour in zero_hours:
            note = "nenhum WAV a começar nesta hora"
        elif n_events == 0:
            note = "sem arquivos com data nesta hora"
        else:
            note = ""
        ws.append([f"{hour:02d}h", n_events, note])
        row = ws.max_row
        fill = PEAK_FILL if hour in peak_hours else ZERO_FILL if hour in zero_hours else None
        if fill:
            for col in range(1, 4):
                ws.cell(row, col).fill = fill
    _bar_chart(
        ws,
        "Eventos por hora (início da gravação)",
        "Hora (início do arquivo)",
        header_row=2,
        n_rows=24,
    )


def _write_monthly_sheet(ws, results: list[PipelineResult]) -> None:
    titles = ["Mês", "Eventos", "Nota"]
    _note_row(
        ws,
        1,
        3,
        "Mês = recorded_at do INÍCIO do arquivo. MP3s sem data no nome não entram neste gráfico.",
    )
    _headers(ws, 2, titles)
    for item in _monthly_counts(results):
        month = int(item["month"])
        n_events = int(item["n_events"])
        label = f"{month:02d} {MONTH_LABELS[month - 1]}"
        if n_events == 0:
            note = "sem arquivos com data neste mês"
            fill = ZERO_FILL
        else:
            note = ""
            fill = None
        ws.append([label, n_events, note])
        row = ws.max_row
        if fill:
            for col in range(1, 4):
                ws.cell(row, col).fill = fill
    _bar_chart(
        ws,
        "Eventos por mês (início da gravação)",
        "Mês (início do arquivo)",
        header_row=2,
        n_rows=12,
    )
